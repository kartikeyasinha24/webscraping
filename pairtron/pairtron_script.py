import re
import pandas as pd
import pandas.io.formats.excel
from configparser import ConfigParser
from datetime import datetime
from termcolor import cprint
import os
import shutil
from openpyxl import load_workbook
import numpy as np

class pairtron():

    def affiliation_cleaner(self, affiliation):
        # print(affiliation)
        affiliation = str(affiliation)
        affiliation = affiliation.strip(" ;").replace("   ", " ").replace("  "," ")
        while ' ;' in affiliation:
            affiliation = affiliation.replace(" ;", ";")
        while ';;' in affiliation:
            affiliation = affiliation.replace(";;", ";")
        return affiliation

    def zeta0_creation(self, indexed_files_dir, merge_columns):
        """ Returns pandas dataframe which has latest record for each manual id after merging all "sheet_name"
        in the previously indexed_files which are present in "indexed_files_dir"
        """
        indexed_files = [file for file in os.listdir(indexed_files_dir) if not file.startswith("~")]

        indexed_files_dict = {}
        indexed_files_dict.clear()

        dateList = []
        del dateList[:]
        for file in indexed_files:
            dated = file.split('_')[-1].split('.')[0]
            dated = dated[4:] + dated[:4]
            dateList.append(dated)
            indexed_files_dict[dated] = file

        dataframes = {}

        for dated, file in indexed_files_dict.items():
            file_name = indexed_files_dir + '\\' + file
            dataframes[dated] = pd.read_excel(file_name, sheet_name=0)
            dataframes[dated]['file_date'] = dated
            dataframes[dated]['mid'] = [int(elem.split('_')[-1]) for elem in dataframes[dated]['manual_id']]

        merged_df = pd.concat([dataframes[dated] for dated in dateList], ignore_index=True)
        merged_df = merged_df.sort_values('file_date', ascending=False)
        zeta0 = merged_df.drop_duplicates(subset='manual_id', keep='first')
        pd.set_option('mode.chained_assignment', None)
        for col in zeta0.columns:
            zeta0[col] = zeta0[col].astype('str')
        zeta0 = zeta0.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        zeta0 = zeta0.sort_values('mid', ascending=True)
        if "manual_id" not in merge_columns:
            merge_columns.append("manual_id")
        zeta0 = zeta0[merge_columns]
        # print(zeta0)
        return zeta0

    def copy_larvol_xlsx(self, template, acronym):
        date = datetime.now().date().strftime('%m%d%Y')
        self.dest_file = os.path.basename(template).replace('ACRONYM',acronym).replace('MMDDYYYY', date + '_Pairtron')
        shutil.copy2(template, self.dest_file)

    def selectionAfterJoin(self, df, cols, common_cols):
        for col in common_cols:
            if col != 'manual_id':
                df[col] = np.where(df['{}_left'.format(col)].isnull() | ((df['{}_right'.format(col)].notnull()) & (df['{}_right'.format(col)] != '') & (df['{}_left'.format(col)] != df['{}_right'.format(col)])), df['{}_right'.format(col)], df['{}_left'.format(col)])
        drop_list = ['{}_left'.format(col) for col in common_cols if col != 'manual_id']
        drop_list.extend(['{}_right'.format(col) for col in common_cols if col != 'manual_id'])
        df.drop(drop_list, axis=1, inplace=True)
        return df[cols]

    def update_larvol_xlsx(self, src, acronym, sheets, columns, zeta0_df=None):
        wb = load_workbook(filename=self.dest_file)
        ws = wb[sheets[0]]
        ws.title = sheets[0].replace('ACRONYM',acronym)
        try:
            curr_df = pd.read_excel(src)
        except:
            curr_df = pd.read_csv(src)
        if zeta0_df is not None:
            curr_jn_zeta = pd.merge(curr_df, zeta0_df, left_on='manual_id', right_on='manual_id', how='left', suffixes=('_left', '_right'))
            common_columns = [col for col in curr_df.columns if col in zeta0_df.columns]
            self.source_df = self.selectionAfterJoin(curr_jn_zeta, columns, common_columns)
        else:
            self.source_df = curr_df
        session_list = self.source_df.fillna('').values.tolist()
        for row_iter in range(len(session_list)):
            print(row_iter)
            for col_iter in range(len(session_list[row_iter])):
                print(col_iter)
                ws.cell(row=row_iter+2, column=col_iter+1).value = self.affiliation_cleaner(session_list[row_iter][col_iter])
        wb.save(self.dest_file)

    def process_merger(self, source_file, acronym, merge, template, sheets, indexed_files_dir, columns, merge_columns):
        self.copy_larvol_xlsx(template, acronym)
        if merge.upper() == 'YES':
            zeta0 = self.zeta0_creation(indexed_files_dir, merge_columns)
            self.update_larvol_xlsx(source_file, acronym, sheets, columns, zeta0)
        else:
            self.update_larvol_xlsx(source_file, acronym, sheets, columns)

    def separated_by_number(self, source_id, manual_id, authors_list, affiliations_list):
        separated_authors_list = []
        affiliations_dict = {}
        prev_affiliation = None

        for affiliation in affiliations_list:
            #print(manual_id)
            #print(affiliation)
            if affiliation != '':
                group = re.findall(r'\d+', affiliation)
                #print(group)
                if group != []:
                    num = list(map(int, group))[0]
                    affiliations_dict[str(num)] = str(num).join(affiliation.split(str(num))[1:]).strip(',. ')
                    prev_affiliation = num
                elif prev_affiliation is not None:
                    num = prev_affiliation
                    affiliations_dict[str(num)] = affiliations_dict[str(num)] + '; ' + affiliation.strip(',. ')
                    prev_affiliation = num

        for author in authors_list:
            #print(author)
            group = re.findall(r'\d+', author)
            num_list = list(map(int, group))
            #print(num_list)
            if num_list != []:
                author_name = author.split(str(num_list[0]))[0].strip(',.-; ')
            else:
                author_name = author.strip(',.-; ')
            #print(author_name)
            for num in num_list:
                try:
                    elem = affiliations_dict[str(num)]
                except:
                    affiliations_dict[str(num)] = ''
                    cprint("Exception for manual_id: {} as affiliation index {} wasn't found".format(manual_id, str(num)), 'yellow', attrs=['bold'])

            affiliation_name = '; '.join([affiliations_dict[str(num)].strip(',.- ') for num in num_list])
            #print(affiliation_name)
            separated_authors_list.append([source_id, manual_id, author_name, affiliation_name])

        return separated_authors_list

    def separated_by_semicolon(self, source_id, manual_id, authors_list, affiliations_list):
        separated_authors_list = []

        for iter in range(len(authors_list)):
            author_name = authors_list[iter].strip(',.-; ')
            try:
                affiliation_name = affiliations_list[iter].strip(',.- ')
            except:
                affiliation_name = ''
            separated_authors_list.append([source_id, manual_id, author_name, affiliation_name])

        return separated_authors_list


    def common_affiliation(self, source_id, manual_id, authors_list, affiliations_list):
        separated_authors_list = []

        for iter in range(len(authors_list)):
            author_name = authors_list[iter].strip(',.-; ')
            affiliation_name = affiliations_list[0].strip(',.- ')
            print(affiliation_name)
            separated_authors_list.append([source_id, manual_id, author_name, affiliation_name])

        return separated_authors_list

    def process_pairtron(self, sheet):
        source_df = self.source_df
        source_df = source_df[source_df['authors'].notnull()]
        source_id_list = source_df['source_id'].fillna('').tolist()
        manual_id_list = source_df['manual_id'].fillna('').tolist()
        authors_list = source_df['authors'].tolist()
        affiliation_list = source_df['author_affiliation'].fillna('').tolist()
        pairtron_list = []

        for iter in range(len(authors_list)):
            #print(iter, manual_id_list[iter])
            author_tokens = [elem.strip() for elem in authors_list[iter].split(';')]
            affiliation_tokens = [elem.strip() for elem in affiliation_list[iter].split(';')]
            try:
                if author_tokens[0][-1].isdigit() and '1' in affiliation_list[iter]:
                    pairtron_list.extend(self.separated_by_number(source_id_list[iter], manual_id_list[iter], author_tokens, affiliation_tokens))
                elif len(author_tokens) == len(affiliation_tokens):
                    pairtron_list.extend(self.separated_by_semicolon(source_id_list[iter], manual_id_list[iter], author_tokens, affiliation_tokens))
                elif author_tokens[0][-1].isdigit() and '1' not in affiliation_list[iter]:
                    cprint("ALERT: manual_id: {} has missing affiliations.".format(manual_id_list[iter]), 'red', attrs=['bold'])
                else:
                    pairtron_list.extend(self.common_affiliation(source_id_list[iter], manual_id_list[iter], author_tokens, affiliation_tokens))
            except:
                pass
        df = pd.DataFrame(pairtron_list, columns=['source_id', 'manual_id', 'authors', 'author_affiliation'])
        df.drop_duplicates(inplace = True)
        authorsInfo_list = df.values.tolist()
        wb = load_workbook(filename=self.dest_file)
        ws = wb[sheet]
        for row_iter in range(len(authorsInfo_list)):
            for col_iter in range(len(authorsInfo_list[row_iter])):
                ws.cell(row=row_iter+2, column=col_iter+1).value = authorsInfo_list[row_iter][col_iter]
        wb.save(self.dest_file)

    def processData(self, source_file, acronym, merge, template, sheets, indexed_files_dir, columns, merge_columns):
        self.process_merger(source_file, acronym, merge, template, sheets, indexed_files_dir, columns, merge_columns)
        self.process_pairtron(sheets[1])

if __name__ == "__main__":

    start = datetime.now()
    print ("Script Start Time ",start)
    print ("Script Running.....\n")

    parser = ConfigParser()
    parser.read('pairtron_config.ini')

    source_file = parser.get('dynamic_fields', 'source_file')
    acronym = parser.get('dynamic_fields', 'ACRONYM')
    merge = parser.get('dynamic_fields', 'merge')
    merge_columns = [elem.strip() for elem in parser.get('dynamic_fields', 'merge_columns').split(',')]
    template = parser.get('static_fields', 'template')
    sheets = parser.get('static_fields', 'sheets').split(',')
    indexed_files_dir = parser.get('static_fields', 'indexed_files_dir')
    columns = parser.get('static_fields', 'columns').split(',')

    obj = pairtron()
    obj.processData(source_file, acronym, merge, template, sheets, indexed_files_dir, columns, merge_columns)

    total_time = datetime.now() - start
    print ("\nScript End Time ",datetime.now())
    print ("Execution Time", total_time)